import openpyxl, sys #pyperclip?


# def iterate(current_sheet, team_1, team_2):
#     for j in range(3, 355):
#         sheet = current_sheet
#         global team_1_value
#         global team_2_value

            
#         if team_1 == str(sheet.cell(row=j, column=1).value):
#             #print(str(team_1) + ': ' + str(sheet.cell(row=j, column=2).value))
#             team_1_value += float(sheet.cell(row=j, column=2).value)
#            # print(team_1_value)
                    
#         if team_2 == str(sheet.cell(row=j, column=1).value):
#             #print(str(team_2) + ': ' + str(sheet.cell(row=j, column=2).value))
#             team_2_value += float(sheet.cell(row=j, column=2).value)
#            # print(team_2_value)
    
team_1_value = 0
team_2_value = 0
        
while True:
    team_1_value = 0
    team_2_value = 0
    try:
        
        print('Enter team #1: ', end='')
        team_1 = input()
        print('Enter team #1 seed (Lower seed): ', end='')
        team_1_seed = input()
        print('Enter team #2: ', end='')
        team_2 = input()
        print('Enter team #2 seed (Higher seed): ', end=' ')
        team_2_seed = input()
        print(' ')
        wb = openpyxl.load_workbook('example2_copy.xlsx')
        #print(str(wb.sheetnames))
        seed_difference = int(team_2_seed) - int(team_1_seed)
        sheet = wb['Offensive Efficiency']


        # for i in wb.sheetnames:
        #     sheet = wb[i]
        #    #iterate(wb[i], team_1, team_2)
        #     print(str(wb[i]))
        #     for j in range(3, 355):
        #         if team1 == str(sheet.cell(row=j, column=1).value):
        #            # print(str(team1) + ': ' + str(sheet.cell(row=j, column=2).value))
        #             team_1_value += float(sheet.cell(row=j, column=2).value)
                    
        #         if team2 == str(sheet.cell(row=j, column=1).value):
        #            # print(str(team2) + ': ' + str(sheet.cell(row=j, column=2).value))
        #             team_2_value += float(sheet.cell(row=j, column=2).value)
        
        for i in wb.sheetnames:
            sheet = wb[wb[i].title]
            #print(sheet.title) 

            for j in range(3, 355):

                if team_1 == str(sheet.cell(row=j, column=1).value):
                    if wb[i].title == 'Offensive Efficiency':
                        team_1_value += float(sheet.cell(row=j, column=2).value * 0.4)
                        print(team_1 + ' Offensive Efficiency: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Defensive Efficiency':
                        team_1_value -= float(sheet.cell(row=j, column=2).value * 0.35)
                        print(team_1 + ' Defensive Efficiency: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'eFG %':
                        team_1_value += float(sheet.cell(row=j, column=2).value * 0.4)
                        print(team_1 + ' eFG %: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Opponent eFG %':
                        team_1_value -= float(sheet.cell(row=j, column=2).value * 0.35)
                        print(team_1 + ' Opponent eFG %: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'TS %':
                        team_1_value += float(sheet.cell(row=j, column=2).value * 0.4)
                        print(team_1 + ' TS %: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Opponent TS %':
                        team_1_value -= float(sheet.cell(row=j, column=2).value * 0.35)
                        print(team_1 + ' Opponent TS %: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Offensive Rebounding %':
                        team_1_value += (float(sheet.cell(row=j, column=2).value) * 0.3)
                        print(team_1 + ' Offensive Rebounding %: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Defensive Rebounding %':
                        team_1_value += (float(sheet.cell(row=j, column=2).value) * 0.25)
                        print(team_1 + ' Defensive Rebounding %: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'TO per POS':
                        team_1_value -= (float(sheet.cell(row=j, column=2).value) * 0.2)
                        print(team_1 + ' TO per POS: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Opponent TO per POS':
                        team_1_value += (float(sheet.cell(row=j, column=2).value) * 0.2) 
                        print(team_1 + ' Opponent TO per POS: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Free Throw %':
                            team_1_value += (float(sheet.cell(row=j, column=2).value) * 0.1)  

                    elif wb[i].title == 'SOS and Win %':
                            team_1_value *= (float(sheet.cell(row=j, column=4).value))  
                        
                if team_2 == str(sheet.cell(row=j, column=1).value):
                    if wb[i].title == 'Offensive Efficiency':
                        team_2_value += float(sheet.cell(row=j, column=2).value * 0.4)
                        print(team_2 + ' Offensive Efficiency: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Defensive Efficiency':
                        team_2_value -= float(sheet.cell(row=j, column=2).value * 0.35)
                        print(team_2 + ' Defensive Efficiency: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'eFG %':
                        team_2_value += float(sheet.cell(row=j, column=2).value * 0.4)
                        print(team_2 + ' eFG %: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Opponent eFG %':
                        team_2_value -= float(sheet.cell(row=j, column=2).value * 0.35)
                        print(team_2 + ' Opponent eFG %: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'TS %':
                        team_2_value += float(sheet.cell(row=j, column=2).value * 0.4)
                        print(team_2 + ' TS %: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Opponent TS %':
                        team_2_value -= float(sheet.cell(row=j, column=2).value * 0.35)
                        print(team_2 + ' Opponent TS %: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Offensive Rebounding %':
                        team_2_value += (float(sheet.cell(row=j, column=2).value) * 0.3)
                        print(team_2 + ' Offensive Rebounding %: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Defensive Rebounding %':
                        team_1_value += (float(sheet.cell(row=j, column=2).value) * 0.25)
                        print(team_2 + ' Defensive Rebounding %: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'TO per POS':
                        team_2_value -= (float(sheet.cell(row=j, column=2).value) * 0.2)
                        print(team_2 + ' TO per POS: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Opponent TO per POS':
                        team_2_value += (float(sheet.cell(row=j, column=2).value) * 0.2) 
                        print(team_2 + ' Opponent TO per POS: ' + str(sheet.cell(row=j, column=2).value))

                    elif wb[i].title == 'Free Throw %':
                            team_2_value += (float(sheet.cell(row=j, column=2).value) * 0.1)  

                    elif wb[i].title == 'SOS and Win %':
                            team_2_value *= (float(sheet.cell(row=j, column=4).value))  

        
        if seed_difference < 3:
            team_1_value *= 1 - (((100/15) * seed_difference ) / 100)
        print(' ')
        if team_1_value > team_2_value and (team_1_value - team_2_value) >= 0.05:
            print(f'Matchup winner: {team_1}')
        else:
            print(f'Matchup winner: {team_2}')
        print(team_1_value)
        print(team_2_value)


    except KeyboardInterrupt:
        sys.exit()


